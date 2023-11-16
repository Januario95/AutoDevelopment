#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
"""
Created on Wed May 17, 2023  13:53:41

@author: a248433
"""

import os
import sys
import json
import time
import random
import logging
import keyring
import traceback
import pandas as pd
from selenium import webdriver
import chromedriver_autoinstaller
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

INDEXES = list(range(1, 1000))

def get_credentials(sys_name, username):
    """
    Fetch credentials from Windows Credentials
    """
    keyring.get_keyring()
    password = keyring.get_password(sys_name, username)
    return password

f = open("config.txt")
configInfo = json.loads(f.read())

input_username = configInfo["input_username"]
input_key_name = configInfo["input_key_name"]
input_password = get_credentials(input_key_name, input_username) #.values()
# print(input_password)

auth_username = configInfo["auth_username"]
auth_key_name = configInfo["auth_key_name"]
auth_password = input_password

T24_URL = configInfo['url']

# print({'input_username': input_username, 'input_password': input_password})
# print({'auth_username': auth_username, 'auth_password': auth_password})

def display_text_std(text, delay=0.000000001):
    """
    Interatively format output in the console
    """
    sys.stdout.write(f'\t')
    for char in text:
        sys.stdout.write(f'{char}')
        sys.stdout.flush()
        time.sleep(delay)

    print('')

def find_input_field_and_fill_value(driver, input_value, text_to_console, x_path):
    global INDEXES
    try:
        tag = driver.find_element(By.XPATH, x_path)
        tag.clear()
        tag.send_keys(input_value)
        # for char in input_value:
        #     tag.send_keys(char)
        #     time.sleep(0.0000001)
        text = f'{INDEXES.pop(0)}. Bot inserted {text_to_console!r} into input field.   {chr(482)}'
        # display_text_std(text)
    except Exception as e:
        text_to_console = e.args
        text = f'{INDEXES.pop(0)}. Bot inserted {text_to_console!r} into input field.   {chr(482)}'
        # display_text_std(text)

class ExtractValue:
    def __init__(self, df):
        self.df = df

    def update_val(self, filename, cell_nr, val):
        # path = os.getcwd() + f'\\Clients\\approved\{foldername}\\data.xlsx'
        # path = os.getcwd() + f'\\data.xlsx'
        path = f'C:/Users/a248433/Documents/SB Mozambique/EDO/Robotics/Bots/Automatização do Processo Abertura de contas Corporate/Clientes/approved/{filename}/data.xlsx'
        # os.chdir(f'./Clientes/approved/{foldername}')
        # file_path = 'data.xlsx'
        wb = load_workbook(path)
        ws = wb['Contas Sociedade Constituida Pr']
        ws[cell_nr] = val
        wb.save(path)
        text = f'{INDEXES.pop(0)}. Bot is updating cell {cell_nr} with value={val}.   {chr(482)}'
        display_text_std(text)

    def get_value(self, key):
        value = None
        for row in self.df.itertuples():
            if row.Index == key:
                value = row.Value
        return value


class ClientesSociedadePrenchido(ExtractValue):
    def __init__(self, filename):
        self.xls = pd.ExcelFile(filename)
        self.df = pd.read_excel(self.xls, sheet_name='Clientes Sociedade Prenchido')
        self.df = self.df[['__________________Informacoes dos Documentos de Identificacao_________________',
            'Unnamed: 1', 'Unnamed: 2']]
        self.df = self.df.iloc[1:94, :]
        self.df = self.df.rename(columns={
        '__________________Informacoes dos Documentos de Identificacao_________________': 'Name',
        'Unnamed: 1': 'Value', 'Unnamed: 2': 'Meaning'
        })
        columns = self.df.columns
        for col in columns:
            self.df[col] = self.df[col].astype(str)
        self.df.set_index('Name', inplace=True)

    
class ContasSociedadeConstituidaPr(ExtractValue):
    def __init__(self, filename):
        self.xls = pd.ExcelFile(filename)
        df = pd.read_excel(self.xls, sheet_name='Contas Sociedade Constituida Pr')
        last_col = df.columns[-1]
        new_row = {
            'Name': df.columns[0],
            'Value': df.columns[-1]
        }
        df = df.rename(columns={
            'No. do Cliente': 'Name',
            last_col: 'Value'
        })
        df = df.iloc[1:21, :]
        df = df.append(new_row, ignore_index=True)
        self.df = df
        columns = self.df.columns
        for col in columns:
            self.df[col] = self.df[col].astype(str)
        self.df.set_index('Name', inplace=True)
    

class ClientCIFs(ExtractValue):
    def __init__(self, filename):
        self.xls = pd.ExcelFile(filename)
        self.df = pd.read_excel(self.xls, sheet_name='CIF_CUS')
    

def progress_bar(iterations):
    for k in range(iterations+1):
        percentage = k / iterations
        time_msg = "\rGenerating letter {0:.2%} ".format(percentage)
        sys.stdout.write(time_msg)
        sys.stdout.flush()
        time.sleep(0.000000001)
    
def get_driver():
    """
    Automatically install chromedrivre is unavailable and
    instantiate a driver instance for Chrome
    """
    chromedriver_autoinstaller.install()
    options = webdriver.ChromeOptions()
    options.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=options)
    return driver


def find_element_by_xpath_and_click(driver, xpath_selector, delay=1):
    """
    Locate an element by XPATH in the page and click in it
    """
    element = driver.find_element(By.XPATH, xpath_selector)
    element.click()
    time.sleep(delay)

def login_user(driver, user_type):
    global input_username, input_password, auth_username, auth_password
    if user_type == 'input_user':
        username = input_username
        password = input_password
        display_text_std(f'{INDEXES.pop(0)}. Logging in as Float.   {chr(482)}')
    elif user_type == 'auth_user':
        username = 'IVOFAQ'
        password = '@AAbb123'
        display_text_std(f'{INDEXES.pop(0)}. Logging in as Team Leader.   {chr(482)}')

    username_field = driver.find_element(By.XPATH, '//*[@id="signOnName"]')
    password_field = driver.find_element(By.XPATH, '//*[@id="password"]')
    username_field.send_keys(username)
    display_text_std(f'{INDEXES.pop(0)}. Bot entered username.   {chr(482)}')
    password_field.send_keys(password)
    display_text_std(f'{INDEXES.pop(0)}. Bot entered password.   {chr(482)}')
    password_field.send_keys(Keys.RETURN)
    time.sleep(3)

def logout_user(driver):
    iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "frame")))
    driver.switch_to.frame(iframe)

    sign_off_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="pane_"]/div[1]/div/table/tbody/tr/td[3]/a')))
    sign_off_btn.click()
    time.sleep(1)

def find_element_by_xpath_and_click(driver, selector, delay=0.6):
    element = driver.find_element(By.XPATH, selector)
    element.click()
    time.sleep(delay)

def common_handler(driver, client_nr):
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "frame")))    
    frames = driver.find_elements(By.TAG_NAME, "frame") 
    frame = frames[1]
    driver.switch_to.frame(frame)
    display_text_std(f'{INDEXES.pop(0)}. Bot switched driver to new frame.   {chr(482)}')

    menu = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pane_"]/ul/li/span')))
    menu.click()
    menu_float = 'Menu Float'
    text = f'{INDEXES.pop(0)}. Bot clicked on {menu_float!r}.   {chr(482)}'
    display_text_std(text)
    time.sleep(1)

    text = 'Cliente'
    client_menu = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pane_"]/ul/li/ul/li[1]/span')))
    client_menu.click()
    text = f'{INDEXES.pop(0)}. Bot clicked on {text!r}.   {chr(482)}'
    display_text_std(text)
    time.sleep(1)

    text = 'Abertura Cliente em Nome Individual'
    find_element_by_xpath_and_click(driver, '//*[@id="pane_"]/ul/li/ul/li[1]/ul/li[1]/a', delay=3)
    text = f'{INDEXES.pop(0)}. Bot clicked on {text!r}.   {chr(482)}'
    display_text_std(text)

    window_handles = driver.window_handles
    size = len(window_handles)
    another_window = window_handles[size-1]
    driver.switch_to.window(another_window)
    driver.maximize_window()
    text = f'{INDEXES.pop(0)}. Bot switched driver to a new window.   {chr(482)}'
    display_text_std(text)
    time.sleep(1)

    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/form[1]")))    

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div[2]/form[1]/div[2]/table/thead/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/input[1]')))
    search_field = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/form[1]/div[2]/table/thead/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/input[1]')
    search_field.send_keys(client_nr)
    text = f'{INDEXES.pop(0)}. Bot inserted client_nr={client_nr!r} into field.   {chr(482)}'
    display_text_std(text)
    time.sleep(3)

def change_to_last_window(driver):
    global INDEXES
    text = f'{INDEXES.pop(0)}. Bot is attempting to change window.   {chr(482)}'
    display_text_std(text)
    window_handles = driver.window_handles
    last_window = window_handles[-1]
    driver.switch_to.window(last_window)
    driver.maximize_window()
    text = f'{INDEXES.pop(0)}. Bot successfully changed window.   {chr(482)}'
    display_text_std(text)
    time.sleep(2)

def handle_user_auth(driver, client_nr):
    # Handle user authorization
    login_user(driver, 'auth_user')

    common_handler(driver, client_nr)

    find_element_by_xpath_and_click(
        driver,
        selector='/html/body/div[3]/div[2]/form[1]/div[2]/table/thead/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/div/table/tbody/tr/td[3]/a/img',
        delay=1)
    text = f'{INDEXES.pop(0)}. Bot clicked on perform_action_on_contract.   {chr(482)}'
    display_text_std(text)

    try:
        authorize_deal = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/form[1]/div[2]/table/thead/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/div/table/tbody/tr/td[5]/img')
        authorize_deal.click()
        text = f'{INDEXES.pop(0)}. The bot successfully validated the client inhibitation from using new checks.   {chr(482)}'
        display_text_std(text)
        time.sleep(2)
    except Exception as e:
        text = f'{INDEXES.pop(0)}. Unable to validate deal.   {chr(530)}'
        display_text_std(text)

def change_branch(driver, branch_name):
    tools_field = driver.find_element(By.XPATH, '/html/body/div[3]/div[1]/div/table/tbody/tr/td[2]/a')
    tools_field.click()
    text = 'Tools'
    text = f'{INDEXES.pop(0)}. Bot clicked on {text!r}.   {chr(482)}'
    display_text_std(text)
    change_to_last_window(driver)
    text = f'{INDEXES.pop(0)}. Bot switched to a new tab.   {chr(482)}'
    display_text_std(text)
    time.sleep(3)

    companies_field = driver.find_element(By.XPATH, '//*[@id="pane_"]/ul[2]/li/span')
    companies_field.click()
    text = 'Companies'
    text = f'{INDEXES.pop(0)}. Bot clicked on {text!r}.   {chr(482)}'
    display_text_std(text)

    branch_button = driver.find_element(By.LINK_TEXT, branch_name)
    branch_button.click()
    text = f'{INDEXES.pop(0)}. Bot switched to branch = {branch_name!r}.   {chr(482)}'
    display_text_std(text)
    time.sleep(2)