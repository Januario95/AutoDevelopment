import os
import sys
import json
import time
import pandas as pd
from openpyxl import load_workbook


class ExtractValue:
    def __init__(self, df):
        self.df = df

    def update_val(self, cell_nr, val):
        file_path = 'data2.xlsx'
        wb = load_workbook(file_path)
        ws = wb['Contas Sociedade Constituida Pr']
        ws[cell_nr] = val
        wb.save(file_path)


    def get_value(self, key):
        value = None
        for row in self.df.itertuples():
            if row.Index == key:
                value = row.Value
        return value

class ClientesSociedadePrenchido(ExtractValue):
    def __init__(self, filename='data2.xlsx'):
        self.xls = pd.ExcelFile(filename)
        self.df = pd.read_excel(self.xls, sheet_name='Clientes Sociedade Prenchido')
        self.df = self.df[['__________________Informacoes dos Documentos de Identificacao_________________',
            'Unnamed: 1', 'Unnamed: 2']]
        self.df = self.df.iloc[1:94, :]
        self.df = self.df.rename(columns={
        '__________________Informacoes dos Documentos de Identificacao_________________': 'Name',
        'Unnamed: 1': 'Value', 'Unnamed: 2': 'Meaning'
        })
        columns = self.df.columns
        for col in columns:
            self.df[col] = self.df[col].astype(str)
        self.df.set_index('Name', inplace=True)

    
class ContasSociedadeConstituidaPr(ExtractValue):
    def __init__(self, filename='data2.xlsx'):
        self.xls = pd.ExcelFile(filename)
        df = pd.read_excel(self.xls, sheet_name='Contas Sociedade Constituida Pr')
        last_col = df.columns[-1]
        new_row = {
            'Name': df.columns[0],
            'Value': df.columns[-1]
        }
        df = df.rename(columns={
            'No. do Cliente': 'Name',
            last_col: 'Value'
        })
        df = df.iloc[1:21, :]
        df = df.append(new_row, ignore_index=True)
        self.df = df
        columns = self.df.columns
        for col in columns:
            self.df[col] = self.df[col].astype(str)
        self.df.set_index('Name', inplace=True)

class ClientCIFs(ExtractValue):
    def __init__(self, filename='data2.xlsx'):
        self.xls = pd.ExcelFile(filename)
        self.df = pd.read_excel(self.xls, sheet_name='CIF_CUS')

# client_cif_df = ClientCIFs()
# print(client_cif_df.df)

# client_sociedade = ClientesSociedadePrenchido()
# print(client_sociedade.get_value('Customer Since'))
# print(client_sociedade.get_value('N.Contribuinte (NUIT).1'))
# print(client_sociedade.df1.dtypes)
# # print(df.extract_value())
# print(client_sociedade.df)
# print(client_sociedade.get_value('GB # Nome:'))

# sociedade_constituida = ContasSociedadeConstituidaPr()
# print(sociedade_constituida.df)
# sociedade_constituida.update_val('B1', '1111')
# print(sociedade_constituida.get_value('No. do Cliente'))
# print(sociedade_constituida.df)
# print(sociedade_constituida.get_value('No. do Cliente'))
# print(sociedade_constituida.get_value('GB # Nome da Conta'))


# df = pd.read_excel('data2.xlsx', sheet_name='CIF_CUS')
# print(df)
# index = 15
# counter = 0
# for row in client_cif_df.df.itertuples():
#     NOME_da_PESSOA = row.NOME
#     CARGO_NA_EMPRESA = row.CARGO
#     NUMERO_DO_CLIENTE = str(row.CIF)

